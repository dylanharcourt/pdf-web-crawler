from collections import defaultdict
from bs4 import BeautifulSoup
from openpyxl.worksheet.worksheet import Worksheet
from requests.models import HTTPError, Response
from typing import List, Tuple
from urllib.parse import urljoin
from urllib.error import URLError

import fitz
import openpyxl
import requests
import re
import os


class PdfBot:
    """
    PdfBot represents a web scraper that can go to a webpage and download PDFs
    """

    def __init__(self, base_url=""):
        if base_url:
            self.pdf_urls = []
            self.base_url = base_url
            self.soup = BeautifulSoup(self._url_request(base_url).text, "html.parser")

    def extract_data_from_pdfs(self) -> Tuple[str, fitz.Document]:
        """
        Extract pdf content from the list of pdf_urls and yield each one
        :return: Tuple[str, fitz.Document], the requested url and pdf content
        """
        for url in self.pdf_urls:
            yield from self.extract_data_from_pdf(url)

    def extract_data_from_pdf(self, url: str) -> Tuple[str, fitz.Document]:
        """
        Extract pdf content from a single url and yield the url and document
        :param url: str,
        :return: Tuple[str, fitz.Document], the requested url and pdf content
        """
        req = self._url_request(url)
        with fitz.open(stream=req.content, filetype="pdf") as doc:
            yield url, doc

    def find_matches_in_pdf(self, pdf: fitz.Document, process_map: dict) -> dict:
        """
        Find matching text using regex patterns in pdf document.
        Iterates through regex and pdf pages and returns matches as a dictionary.
        :param pdf: fitz Document, pdf document
        :param process_map: dict, keys are regex (list of regex expressions)
        and pages (optional; must be a valid range)
        :return: dict, list of text matches found for each regex
        """
        matches = defaultdict(list)
        for regex in process_map["regex"]:
            for page in pdf.pages(*process_map.get("pages", None)):
                matches[regex].extend(re.findall(regex, page.get_text()))
        return matches

    def write_to_excel(
        self,
        data: List[List[str]],
        file_path: str,
        header: dict,
        worksheet=None,
        col_wise=False,
        col_start=None,
        row_start=None,
    ) -> None:
        """
        Writes rows or columns to an Excel workbook worksheet
        :param data: List[List[str]], rows or cols to write (should not include header)
        :param file_path: str, the file path to write this Excel file too. Can be relative or absolute. Must include filename.
        :param header: List[str], header row
        :param worksheet (optional): str, name of the worksheet to write to. If it doesn't exist, it will be created
        at the end of the workbook; default: None
        :param col_wise (optional): bool, flag to indicate type of insertion (row or col); default is False (row)
        :param col_start (optional): int, col position to start at; default is None
        :param row_start (optional): int, row position to start at; default is None
        :return: None
        """
        if os.path.exists(file_path):
            wb = openpyxl.load_workbook(file_path)
        else:
            wb = openpyxl.Workbook()

        # set ws
        if worksheet is None:
            ws = wb.active
        else:
            if worksheet not in wb.sheetnames:
                wb.create_sheet(worksheet)
            ws = wb[worksheet]

        # sheet is empty, insert initial data
        if ws.max_row == ws.min_row:
            ws.append(header)

        # append values (row or col wise)
        if col_wise:
            self._append_cols(data, ws, col_start, row_start)
        else:
            for row in data:
                ws.append(row)

        # save ws
        wb.save(file_path)

    def download_pdfs(self, directory: str) -> None:
        """
        Download pdfs to a given directory (will create it, if it does not already exist)
        :param directory: the directory you want to save this file in
        :return: None
        """
        if not (os.path.exists(directory)):
            os.makedirs(directory)
        for url in self.pdf_urls:
            req = self._url_request(url)
            filename = url.split("/")[-1]
            file_path = os.path.join(directory, filename)
            with open(file_path, "wb") as file:
                file.write(req.content)

    def parse_html_for_pdfs(self, regex=r"(.pdf)") -> None:
        """
        Parse the html for PDFs
        :param regex (optional): regex expression to filter PDFs
        :return: None
        """
        og_url = self.soup.find("meta", property="og:url")
        print(regex)
        for pdf_url in self.soup.find_all("a", href=re.compile(regex)):
            print(pdf_url)
            self.pdf_urls.append(urljoin(og_url["content"], pdf_url["href"])) if og_url else self.pdf_urls.append(
                urljoin(self.base_url, pdf_url["href"])
            )

    def _url_request(self, url: str) -> Response:
        """
        Make a request to a url and handle errors appropriately.
        :param url: str, the url we are making the request to
        :return: Response, the HTTP response object
        """
        try:
            return requests.get(url)
        except HTTPError as e:
            print("The server couldn't fulfill the request.")
            print("Error code: ", e.code)

        except URLError as e:
            print("We failed to reach a server.")
            print("Reason: ", e.reason)

        except ValueError as e:
            print("A value error occurred. Please try again.")
            print(e)

        except Exception as e:
            print("An error occurred. Please try again.")
            print(e)

    def _append_cols(self, data: List[List[str]], ws: Worksheet, col_start=None, row_start=None) -> None:
        """
        Insert cells into worksheet starting at row, col position. Inserts column wise
        :param data: List[List[str]], data to insert
        :param was: Worksheet, worksheet to insert into
        :param col_start (optional): int, col position to start at; default is Worksheet.max_column
        :param row_start (optional): int, row position to start at; default is 1
        """
        col_start = ws.max_column if col_start is None else col_start
        row_start = 1 if row_start is None else row_start
        for j, col in enumerate(data):
            for i, val in enumerate(col):
                ws.cell(column=col_start + j + 1, row=i + row_start, value=val)
